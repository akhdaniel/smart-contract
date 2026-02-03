from odoo import models, fields, api
from datetime import datetime
import io
import base64


class ExportRkapSarlog(models.TransientModel):
    _name = 'wizard.export.rkap.sarlog'
    _description = 'Export Realisasi RKAP Sarlog'

    year = fields.Selection(
        selection='_get_available_years',
        string='Tahun',
        required=True
    )

    @api.model
    def _get_available_years(self):
        """Generate selection of years (current year minus 5 to current year)"""
        current_year = datetime.now().year
        years = [(str(year), str(year)) for year in range(current_year - 5, current_year + 1)]
        return years

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        res['year'] = str(datetime.now().year)
        return res

    def action_export_excel(self):
        """Generate Excel file dengan struktur RKAP Realisasi Sarlog"""
        self.ensure_one()

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise Exception("openpyxl library not found. Please install it.")

        # Query budget RKAP berdasarkan tahun yang dipilih
        year_filter = int(self.year)
        budget_records = self.env['vit.budget_rkap'].search([
            ('budget_date', '>=', f'{year_filter}-01-01'),
            ('budget_date', '<=', f'{year_filter}-12-31'),
        ], order='tipe_kegiatan, name')
        
        # Get overall percentage from dashboard statistics (same as Dashboard Sarlog)
        dashboard_stats = self.env['vit.budget_rkap'].get_statistics_sarlog(year_filter)
        overall_percentage = dashboard_stats.get('total_summary', {}).get('persen', 0)

        # Group by tipe_kegiatan and include jenis_penugasan and total pagu izin prinsip
        grouped_data = {}
        for rec in budget_records:
            tipe = rec.tipe_kegiatan if rec.tipe_kegiatan else 'Lainnya'
            if tipe not in grouped_data:
                grouped_data[tipe] = []
            # compute total pagu izin prinsip: prefer stored field, otherwise sum related izin_prinsip
            if rec.total_pagu_izin_prinsip:
                pagu = rec.total_pagu_izin_prinsip
            else:
                pagu = sum(rec.izin_prinsip_ids.mapped('total_pagu')) if rec.izin_prinsip_ids else 0
            
            # Get realisasi data from payments (completed stage)
            payments = self.env['vit.payment'].search([
                ('budget_rkap_id', '=', rec.id),
                ('stage_is_done', '=', True),
            ])
            
            realisasi_pso = 0
            realisasi_kom = 0
            total_realisasi = 0
            item_progress = 0
            
            if payments:
                # Sum realisasi payments based on budget's jenis_penugasan
                total_payment_amount = sum(payments.mapped('amount'))
                
                if rec.jenis_penugasan == 'pso':
                    realisasi_pso = total_payment_amount
                elif rec.jenis_penugasan == 'kom':
                    realisasi_kom = total_payment_amount
                
                total_realisasi = total_payment_amount
                
                # Calculate percentage based on pagu (same as dashboard)
                item_progress = (total_realisasi / pagu * 100) if pagu > 0 else 0
            
            grouped_data[tipe].append({
                'name': rec.name,
                'amount': rec.amount or 0,
                'jenis_penugasan': rec.jenis_penugasan,
                'total_pagu_izin_prinsip': pagu,
                'realisasi_pso': realisasi_pso,
                'realisasi_kom': realisasi_kom,
                'total_realisasi': total_realisasi,
                'avg_progress': item_progress,
            })

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'RKAP Realisasi'

        # Style definitions
        title_font = Font(name='Calibri', size=14, bold=True)
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
        category_font = Font(name='Calibri', size=11, bold=True)
        category_fill = PatternFill(start_color='C5E0B4', end_color='C5E0B4', fill_type='solid')
        subtotal_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        total_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50

        # Row 1: Title
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = f'Rekap Realisasi RKAP Sarlog ({self.year})'
        title_cell.font = Font(name='Calibri', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 25

        # Row 2: Empty
        ws.row_dimensions[2].height = 10

        # Row 3: Grouped header (add RKAP above PSO/KOM/TOTAL)
        group_row = 3
        subheader_row = 4
        headers = ['NO', 'KEGIATAN', 'PSO (Rp)', 'KOM (Rp)', 'TOTAL (Rp)']
        col_letters = ['A', 'B', 'C', 'D', 'E']


        # NO header (merge vertically)
        ws.merge_cells('A3:A4')
        cell = ws['A3']
        cell.value = 'NO'
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

        # KEGIATAN header (merge vertically)
        ws.merge_cells('B3:B4')
        cell = ws['B3']
        cell.value = 'KEGIATAN'
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


        # Group label for RKAP spanning PSO..TOTAL (right side)
        ws.merge_cells('C3:E3')
        grp_cell = ws['C3']
        grp_cell.value = 'RKAP'
        grp_cell.font = header_font
        grp_cell.fill = header_fill
        grp_cell.alignment = center_align
        grp_cell.border = thin_border

        # Group label for REALISASI spanning PSO..PROGRES (far right)
        ws.merge_cells('F3:I3')
        real_cell = ws['F3']
        real_cell.value = 'REALISASI'
        real_cell.font = header_font
        real_cell.fill = header_fill
        real_cell.alignment = center_align
        real_cell.border = thin_border
        ws.row_dimensions[group_row].height = 18

        # SALDO header should be a single merged column (aligned like NO/KEGIATAN)
        ws.merge_cells('J3:J4')
        saldo_cell = ws['J3']
        saldo_cell.value = 'SALDO (Rp)'
        saldo_cell.font = header_font
        saldo_cell.fill = header_fill
        saldo_cell.alignment = center_align
        saldo_cell.border = thin_border

        sub_headers = {
            'C': 'PSO (Rp)',
            'D': 'KOM (Rp)',
            'E': 'TOTAL (Rp)',
            'F': 'PSO (Rp)',
            'G': 'KOM (Rp)',
            'H': 'TOTAL (Rp)',
            'I': 'PROGRES (%)',
        }

        for col, text in sub_headers.items():
            cell = ws[f'{col}4']
            cell.value = text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        ws.row_dimensions[subheader_row].height = 20

        # Set additional width for SALDO column
        ws.column_dimensions['J'].width = 18

        # Set additional column widths
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 15

        # Data rows
        row = 5
        no = 1
        subtotal_rows = {}  # Track subtotal row for each category

        # Mapping nama tipe untuk display yang lebih rapi
        tipe_display_map = {
            'biaya': 'BIAYA',
            'investasi': 'INVESTASI'
        }

        # Sort categories: biaya first, then investasi
        sorted_categories = sorted(grouped_data.keys(), key=lambda x: (0 if x == 'biaya' else 1, x))

        for idx, category in enumerate(sorted_categories):
            items = grouped_data[category]
            category_letter = chr(65 + idx)  # A, B, C, etc
            display_name = tipe_display_map.get(category, category.upper())

            # Category header (e.g., "A BIAYA") - keep NO and KEGIATAN in separate columns
            # choose header fill to match category block: green for biaya, blue for investasi
            right_fill = category_fill if category == 'biaya' else subtotal_fill

            ws[f'A{row}'].value = category_letter
            ws[f'A{row}'].font = category_font
            ws[f'A{row}'].fill = right_fill
            ws[f'A{row}'].alignment = center_align
            ws[f'A{row}'].border = thin_border

            ws[f'B{row}'].value = display_name
            ws[f'B{row}'].font = category_font
            ws[f'B{row}'].fill = right_fill
            ws[f'B{row}'].alignment = left_align
            ws[f'B{row}'].border = thin_border
            ws.row_dimensions[row].height = 18
            # fill right area (C..J) with same category color so header row matches
            for col in ['C','D','E','F','G','H','I','J']:
                ws[f'{col}{row}'].fill = right_fill
                ws[f'{col}{row}'].border = thin_border
                if col in ['C','D','E','F','G','H','J']:
                    ws[f'{col}{row}'].number_format = '#,##0'

            row += 1

            # Items under this category
            category_start_row = row
            category_subtotal = 0
            category_pso = 0
            category_kom = 0
            category_real_pso = 0
            category_real_kom = 0
            category_real_total = 0

            for item in items:
                # Use pagu izin prinsip value to populate PSO or KOM depending on jenis_penugasan
                pagu_val = item.get('total_pagu_izin_prinsip', 0)
                pso_val = pagu_val if item.get('jenis_penugasan') == 'pso' else 0
                kom_val = pagu_val if item.get('jenis_penugasan') == 'kom' else 0
                # total is sum of PSO and KOM (which reflects pagu)
                total_val = pso_val + kom_val

                # Realisasi data
                real_pso = item.get('realisasi_pso', 0)
                real_kom = item.get('realisasi_kom', 0)
                real_total = item.get('total_realisasi', 0)
                progress = item.get('avg_progress', 0)

                ws[f'A{row}'].value = no
                ws[f'B{row}'].value = item['name']
                ws[f'C{row}'].value = pso_val
                ws[f'D{row}'].value = kom_val
                ws[f'E{row}'].value = total_val
                ws[f'F{row}'].value = real_pso
                ws[f'G{row}'].value = real_kom
                ws[f'H{row}'].value = real_total
                ws[f'I{row}'].value = progress

                # Calculate SALDO = TOTAL RKAP - TOTAL REALISASI
                saldo = total_val - real_total
                ws[f'J{row}'].value = saldo

                # choose fill for data area to the right based on category
                data_fill = None
                if category == 'biaya':
                    data_fill = category_fill
                elif category == 'investasi':
                    data_fill = subtotal_fill

                all_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
                for col in all_cols:
                    cell = ws[f'{col}{row}']
                    cell.border = thin_border
                    # apply colored background for entire data area (A..J) when category color present
                    if data_fill:
                        cell.fill = data_fill
                    if col == 'A':
                        cell.alignment = center_align
                    elif col == 'B':
                        cell.alignment = left_align
                    elif col == 'I':
                        cell.alignment = right_align
                        cell.number_format = '0.00'
                    elif col == 'J':
                        cell.alignment = right_align
                        cell.number_format = '#,##0'
                    else:
                        cell.alignment = right_align
                        cell.number_format = '#,##0'

                category_subtotal += total_val
                category_pso += pso_val
                category_kom += kom_val
                category_real_pso += real_pso
                category_real_kom += real_kom
                category_real_total += real_total
                row += 1
                no += 1

            # Subtotal for this category
            ws[f'A{row}'].value = ''
            ws[f'B{row}'].value = f'JUMLAH {display_name}'
            ws[f'B{row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'B{row}'].fill = subtotal_fill
            # write subtotal numbers for PSO, KOM, TOTAL
            ws[f'C{row}'].value = category_pso
            ws[f'D{row}'].value = category_kom
            ws[f'E{row}'].value = category_subtotal
            ws[f'F{row}'].value = category_real_pso
            ws[f'G{row}'].value = category_real_kom
            ws[f'H{row}'].value = category_real_total
            
            # Calculate percentage for category subtotal
            category_percentage = (category_real_total / category_subtotal * 100) if category_subtotal > 0 else 0
            ws[f'I{row}'].value = category_percentage
            ws[f'I{row}'].number_format = '0.00'
            
            # Calculate SALDO for category
            category_saldo = category_subtotal - category_real_total
            ws[f'J{row}'].value = category_saldo
            ws[f'J{row}'].number_format = '#,##0'
            
            all_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
            # choose fill for subtotal area: green for biaya, blue for investasi
            if category == 'biaya':
                right_fill = category_fill
            else:
                right_fill = subtotal_fill
            for col in all_cols:
                ws[f'{col}{row}'].border = thin_border
                # fill both left label area (A,B) and right data area per category
                if col in ['A','B']:
                    ws[f'{col}{row}'].fill = right_fill
                elif col in ['C','D','E','F','G','H','I','J']:
                    ws[f'{col}{row}'].fill = right_fill
                if col in ['C','D','E','F','G','H','J']:
                    ws[f'{col}{row}'].number_format = '#,##0'
            ws[f'B{row}'].alignment = left_align
            ws.row_dimensions[row].height = 16
            
            # Store subtotal row for this category
            subtotal_rows[category] = row
            
            row += 1

        # Grand total row (sum only JUMLAH BIAYA + JUMLAH INVESTASI, not all detail rows)
        ws['A' + str(row)].value = ''
        ws['B' + str(row)].value = 'TOTAL'
        ws['B' + str(row)].font = Font(name='Calibri', size=11, bold=True)
        ws['B' + str(row)].fill = total_fill
        
        # Build formulas that reference only subtotal rows (JUMLAH BIAYA + JUMLAH INVESTASI)
        all_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
        
        # Get rows for jumlah biaya and jumlah investasi
        biaya_row = subtotal_rows.get('biaya')
        investasi_row = subtotal_rows.get('investasi')
        
        if biaya_row and investasi_row:
            # Formula: sum only JUMLAH BIAYA and JUMLAH INVESTASI rows
            for col in ['C', 'D', 'E', 'F', 'G', 'H', 'J']:
                ws[f'{col}{row}'].value = f"={col}{biaya_row}+{col}{investasi_row}"
                ws[f'{col}{row}'].number_format = '#,##0'
        elif biaya_row:
            # Only biaya exists
            for col in ['C', 'D', 'E', 'F', 'G', 'H', 'J']:
                ws[f'{col}{row}'].value = f"={col}{biaya_row}"
                ws[f'{col}{row}'].number_format = '#,##0'
        elif investasi_row:
            # Only investasi exists
            for col in ['C', 'D', 'E', 'F', 'G', 'H', 'J']:
                ws[f'{col}{row}'].value = f"={col}{investasi_row}"
                ws[f'{col}{row}'].number_format = '#,##0'
        
        # Add overall percentage to PROGRES column in TOTAL row
        ws[f'I{row}'].value = overall_percentage
        ws[f'I{row}'].number_format = '0.00'
        
        for col in all_cols:
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].fill = total_fill
        ws['B' + str(row)].alignment = left_align
        ws.row_dimensions[row].height = 18

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = output.read()
        output.close()

        # Encode and save as attachment
        data_b64 = base64.b64encode(file_content).decode('utf-8')
        attachment = self.env['ir.attachment'].create({
            'name': f'realisasi_rkap_sarlog_{self.year}.xlsx',
            'type': 'binary',
            'datas': data_b64,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': 'wizard.export.rkap.sarlog',
            'res_id': self.id,
        })

        # Return download URL
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % (attachment.id,),
            'target': 'new',
        }
