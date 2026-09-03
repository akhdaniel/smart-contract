#!/usr/bin/python
# -*- coding: utf-8 -*-

import base64
import io
import re

from odoo import fields, models, _
from odoo.exceptions import UserError

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class WizardImportKompleksPergudangan(models.TransientModel):
    _name = "wizard.import.kompleks.pergudangan"
    _description = "Wizard Import Kompleks Pergudangan dari Excel"

    _KANWIL_ALIASES = {
        "sumut": ["Sumatera Utara"],
        "sumbar": ["Sumatera Barat"],
        "sumsel": ["Sumatera Selatan"],
        "jabar": ["Jawa Barat"],
        "jateng": ["Jawa Tengah"],
        "jogja": ["DI Yogyakarta", "Daerah Istimewa Yogyakarta", "Yogyakarta"],
        "jatim": ["Jawa Timur"],
        "kalbar": ["Kalimantan Barat"],
        "kaltim": ["Kalimantan Timur"],
        "kalsel": ["Kalimantan Selatan"],
        "kalteng": ["Kalimantan Tengah"],
        "kaltara": ["Kalimantan Utara"],
        "sulut": ["Sulawesi Utara"],
        "sulteng": ["Sulawesi Tengah"],
        "sultra": ["Sulawesi Tenggara"],
        "sulsel": ["Sulawesi Selatan"],
        "ntb": ["Nusa Tenggara Barat"],
        "ntt": ["Nusa Tenggara Timur"],
        "dki": ["DKI Jakarta", "Jakarta"],
        "kepri": ["Kepulauan Riau"],
        "riau kepri": ["Riau Kepri", "Riau dan Kepulauan Riau", "Riau & Kepulauan Riau"],
    }

    file_data = fields.Binary(string="File Excel", required=True)
    file_name = fields.Char(string="File Name")

    def action_import_excel(self):
        if not load_workbook:
            raise UserError(_("Library openpyxl tidak terinstall. Mohon install library openpyxl."))
        if not self.file_data:
            raise UserError(_("Silakan pilih file Excel terlebih dahulu."))

        workbook = load_workbook(io.BytesIO(base64.b64decode(self.file_data)), data_only=True)
        imported_ids = []
        errors = []

        for sheet in workbook.worksheets:
            try:
                imported_ids.extend(self._import_sheet(sheet))
            except Exception as exc:
                errors.append("%s: %s" % (sheet.title, exc))

        imported_ids = list(dict.fromkeys(imported_ids))
        if imported_ids:
            return {
                "type": "ir.actions.act_window",
                "name": _("Imported Kompleks Pergudangan"),
                "res_model": "vit.kompleks_pergudangan",
                "view_mode": "tree,form",
                "domain": [("id", "in", imported_ids)],
                "context": dict(self.env.context, active_test=False),
            }

        message = _("Tidak ada data kompleks pergudangan yang diimport.")
        if errors:
            message += "\n\n" + "\n".join(errors[:10])
        raise UserError(message)

    def _import_sheet(self, sheet):
        header_row = self._find_header_row(sheet)
        if not header_row:
            return []

        columns = self._detect_columns(sheet, header_row)
        current_kanwil = False
        current_kanca = False
        pending_kompleks = False
        imported_ids = []

        def flush_pending_kompleks():
            nonlocal pending_kompleks
            if not pending_kompleks:
                return
            name = ", ".join(pending_kompleks["lines"])
            kompleks = self._find_or_create_kompleks(
                name,
                pending_kompleks["kanwil"],
                pending_kompleks["kanca"],
            )
            imported_ids.append(kompleks.id)
            pending_kompleks = False

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row_values = [self._clean(sheet.cell(row_idx, col).value) for col in range(1, sheet.max_column + 1)]
            if not any(row_values):
                continue

            kanwil_name = self._extract_kanwil_name(row_values, columns)
            if kanwil_name:
                flush_pending_kompleks()
                current_kanwil = self._find_or_create_kanwil(kanwil_name, sheet.title)
                current_kanca = False

            kanca_name = self._extract_kanca_name(row_values, columns)
            tempat_name = self._cell(row_values, columns["tempat_text"])
            if kanca_name:
                flush_pending_kompleks()
                current_kanca = self._find_or_create_kanca(kanca_name, current_kanwil)
            elif tempat_name and current_kanwil and not current_kanca and not self._looks_like_detail(tempat_name):
                current_kanca = self._find_or_create_kanca(tempat_name, current_kanwil)

            kompleks_name = self._extract_kompleks_name(row_values, columns)
            if kompleks_name:
                flush_pending_kompleks()
                if not current_kanwil:
                    current_kanwil = self._find_or_create_kanwil(sheet.title, sheet.title)
                if not current_kanca and tempat_name:
                    current_kanca = self._find_or_create_kanca(tempat_name, current_kanwil)
                if not current_kanca:
                    continue

                pending_kompleks = {
                    "kanwil": current_kanwil,
                    "kanca": current_kanca,
                    "lines": [kompleks_name],
                }
                continue

            detail_line = self._extract_kompleks_detail(row_values, columns)
            if pending_kompleks and detail_line:
                pending_kompleks["lines"].append(detail_line)
                continue

        flush_pending_kompleks()
        return imported_ids

    def _find_header_row(self, sheet):
        for row_idx in range(1, min(sheet.max_row, 50) + 1):
            text = " ".join(
                self._clean(sheet.cell(row_idx, col).value).lower()
                for col in range(1, sheet.max_column + 1)
            )
            if "kantor wilayah" in text and "kantor cabang" in text and "kompleks pergudangan" in text:
                return row_idx
        return False

    def _detect_columns(self, sheet, header_row):
        columns = {
            "kanwil": self._find_header_column(sheet, header_row, "kantor wilayah"),
            "kanca": self._find_header_column(sheet, header_row, "kantor cabang"),
            "tempat": self._find_header_column(sheet, header_row, "tempat kedudukan"),
            "kompleks": self._find_header_column(sheet, header_row, "kompleks pergudangan"),
        }
        missing = [key for key, col in columns.items() if not col]
        if missing:
            raise UserError(_("Kolom tidak ditemukan: %s") % ", ".join(missing))

        columns["kanwil_text"] = self._next_text_column(sheet, header_row, columns["kanwil"], columns["kanca"])
        columns["kanca_number"] = columns["kanca"]
        columns["kanca_text"] = self._next_text_column(sheet, header_row, columns["kanca"], columns["tempat"])
        columns["tempat_text"] = columns["tempat"]
        columns["kompleks_number"] = columns["kompleks"]
        columns["kompleks_text"] = self._next_text_column(sheet, header_row, columns["kompleks"], sheet.max_column + 1)
        return columns

    def _find_header_column(self, sheet, header_row, label):
        label = label.lower()
        for col in range(1, sheet.max_column + 1):
            value = self._clean(sheet.cell(header_row, col).value).lower()
            if label in value:
                return col
        return False

    def _next_text_column(self, sheet, header_row, start_col, end_col):
        best_col = start_col
        best_score = -1
        for col in range(start_col, end_col):
            score = 0
            for row_idx in range(header_row + 1, min(sheet.max_row, header_row + 35) + 1):
                value = self._clean(sheet.cell(row_idx, col).value)
                if value and not self._is_number_marker(value):
                    score += 1
            if score > best_score:
                best_score = score
                best_col = col
        return best_col

    def _extract_kanwil_name(self, row_values, columns):
        value = self._cell(row_values, columns["kanwil_text"])
        if not value or self._looks_like_detail(value) or self._is_number_marker(value):
            return False
        return value

    def _extract_kanca_name(self, row_values, columns):
        marker = self._cell(row_values, columns["kanca_number"])
        value = self._cell(row_values, columns["kanca_text"])
        if not self._is_number_marker(marker):
            return False
        if not value or self._looks_like_detail(value) or self._is_number_marker(value):
            return False
        return value

    def _extract_kompleks_name(self, row_values, columns):
        marker = self._cell(row_values, columns["kompleks_number"])
        value = self._cell(row_values, columns["kompleks_text"])
        if not self._is_number_marker(marker):
            return False
        if not value or self._looks_like_detail(value) or self._is_number_marker(value):
            return False
        return value

    def _extract_kompleks_detail(self, row_values, columns):
        marker = self._cell(row_values, columns["kompleks_number"])
        value = self._cell(row_values, columns["kompleks_text"])
        if self._is_number_marker(marker):
            return False
        if not value or self._is_number_marker(value):
            return False
        return value

    def _find_or_create_kanwil(self, name, sheet_name=False):
        Kanwil = self.env["vit.kanwil"].sudo()
        candidates = self._kanwil_candidates(name, sheet_name)
        for candidate in candidates:
            kanwil = Kanwil.search([("name", "=", candidate)], limit=1)
            if kanwil:
                return kanwil
        for candidate in candidates:
            kanwil = Kanwil.search([("name", "ilike", candidate)], limit=1)
            if kanwil:
                return kanwil
            kanwil = Kanwil.search([("name", "ilike", "Kanwil " + candidate)], limit=1)
            if kanwil:
                return kanwil
        return Kanwil.create({"name": self._kanwil_create_name(name, sheet_name, candidates)})

    def _kanwil_candidates(self, name, sheet_name=False):
        raw_candidates = []
        for value in (name, sheet_name):
            value = self._clean(value)
            if value:
                raw_candidates.append(value)
                raw_candidates.extend(self._KANWIL_ALIASES.get(self._normalize_name(value), []))

        normalized = {self._normalize_name(value) for value in raw_candidates}
        if {"riau", "riau kepri"} & normalized:
            raw_candidates.extend(["Riau", "Kepulauan Riau"])

        candidates = []
        seen = set()
        for value in raw_candidates:
            key = self._normalize_name(value)
            if key and key not in seen:
                seen.add(key)
                candidates.append(value)
        return candidates

    def _normalize_name(self, value):
        value = self._clean(value).lower()
        value = value.replace("&", " dan ")
        value = re.sub(r"\bkanwil\b", " ", value)
        value = re.sub(r"[^a-z0-9]+", " ", value)
        return " ".join(value.split())

    def _kanwil_create_name(self, name, sheet_name, candidates):
        for value in (name, sheet_name):
            aliases = self._KANWIL_ALIASES.get(self._normalize_name(value), [])
            if aliases:
                return aliases[0]
        return candidates[0] if candidates else name

    def _find_or_create_kanca(self, name, kanwil):
        Kanca = self.env["vit.kanca"].sudo()
        domain = [("name", "=", name)]
        if kanwil:
            domain.append(("kanwil_id", "=", kanwil.id))
        kanca = Kanca.search(domain, limit=1)
        if not kanca:
            kanca = Kanca.search([("name", "ilike", name)], limit=1)
        if kanca:
            if kanwil and not kanca.kanwil_id:
                kanca.write({"kanwil_id": kanwil.id})
            return kanca
        return Kanca.create({"name": name, "kanwil_id": kanwil.id if kanwil else False})

    def _find_or_create_kompleks(self, name, kanwil, kanca):
        Kompleks = self.env["vit.kompleks_pergudangan"].sudo()
        short_name = name.split(",", 1)[0]
        kompleks = Kompleks.search([
            ("name", "=", name),
            ("kanwil_id", "in", kanwil.ids),
            ("kanca_id", "in", kanca.ids),
        ], limit=1)
        if not kompleks and short_name != name:
            kompleks = Kompleks.search([
                ("name", "=", short_name),
                ("kanwil_id", "in", kanwil.ids),
                ("kanca_id", "in", kanca.ids),
            ], limit=1)
        values = {
            "name": name,
            "kanwil_id": [(4, kanwil.id)],
            "kanca_id": [(4, kanca.id)],
        }
        if kompleks:
            kompleks.write(values)
            return kompleks
        return Kompleks.create(values)

    def _clean(self, value):
        if value is None:
            return ""
        value = str(value).replace("\xa0", " ").strip()
        return " ".join(value.split())

    def _cell(self, row_values, col):
        if not col or col - 1 >= len(row_values):
            return ""
        return row_values[col - 1]

    def _is_number_marker(self, value):
        return bool(value and re.match(r"^\d+[a-z]?\s*\.?$", str(value).strip(), re.I))

    def _looks_like_detail(self, value):
        value = (value or "").lower().strip()
        return value.startswith((
            "jl",
            "jalan",
            "ds.",
            "ds ",
            "desa",
            "kec",
            "kab",
            "kota",
            "kel",
            "provinsi",
        ))
