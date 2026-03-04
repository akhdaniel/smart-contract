from odoo import models, fields, api
from datetime import datetime
import io
import base64


class RekapKetersediaanSarlog(models.TransientModel):
    _name = 'wizard.rekap.ketersediaan.sarlog'
    _description = 'Wizard Rekap Ketersediaan Sarlog'

    year = fields.Selection(selection='_get_available_years', string='Tahun', required=True)

    @api.model
    def _get_available_years(self):
        current_year = datetime.now().year
        years = [(str(year), str(year)) for year in range(current_year - 5, current_year + 1)]
        return years

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        res['year'] = str(datetime.now().year)
        return res

    def action_export_excel(self):
        """Generate Excel file dengan struktur Rekap Ketersediaan Sarlog"""
        self.ensure_one()

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise Exception("openpyxl library not found. Please install it.")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Rekap Ketersediaan'

        # Style definitions (sama seperti export_rkap_sarlog)
        title_font = Font(name='Calibri', size=14, bold=True)
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
        category_font = Font(name='Calibri', size=11, bold=True)
        category_fill = PatternFill(start_color='C5E0B4', end_color='C5E0B4', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Set column widths for simplified layout
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 18
        # make D and E wider as requested
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 18

        # Row 1: Title spanning A-F
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f'Rekap Ketersediaan RKAP Sarlog ({self.year})'
        title_cell.font = Font(name='Calibri', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 25

        # Row 2: Empty spacer
        ws.row_dimensions[2].height = 10

        # Row 3-4: Headers with group names and "( Rp )" subheaders
        group_row = 3
        subheader_row = 4

        # NO header (merge vertically rows 3-4)
        ws.merge_cells('A3:A4')
        cell = ws['A3']
        cell.value = 'NO'
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

        # KEGIATAN header (merge vertically)
        ws.merge_cells('B3:B4')
        cell = ws['B3']
        cell.value = 'KEGIATAN'
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

        # RKAP group label
        ws.merge_cells('C3:C3')
        grp_cell = ws['C3']
        grp_cell.value = 'RKAP'
        grp_cell.font = header_font
        grp_cell.fill = header_fill
        grp_cell.alignment = center_align
        grp_cell.border = thin_border

        # SISA PEMBAYARAN 2024 group label
        ws.merge_cells('D3:D3')
        cell = ws['D3']
        cell.value = 'SISA PEMBAYARAN 2024'
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

        # IZIN PRINSIP TERBIT 2025 group label
        ws.merge_cells('E3:E3')
        cell = ws['E3']
        cell.value = 'IZIN PRINSIP TERBIT 2025'
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

        # SALDO header (only row 3, subheader will be on row 4)
        saldo_cell = ws['F3']
        saldo_cell.value = 'SISA SALDO'
        saldo_cell.font = header_font
        saldo_cell.fill = header_fill
        saldo_cell.alignment = center_align
        saldo_cell.border = thin_border

        ws.row_dimensions[group_row].height = 18

        # Subheaders below numeric columns
        sub_headers = {
            'C': '( Rp )',
            'D': '( Rp )',
            'E': '( Rp )',
            'F': '( Rp )',
        }
        for col, text in sub_headers.items():
            cell = ws[f'{col}4']
            cell.value = text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        ws.row_dimensions[subheader_row].height = 20

        # Data rows: mimic export_rkap_sarlog structure but with new columns
        row = 5
        no = 1
        subtotal_rows = {}

        # grouping same as earlier (tipe_kegiatan)
        tipe_display_map = {
            'biaya': 'BIAYA',
            'investasi': 'INVESTASI'
        }
        # fetch statistics just like export version; user can adapt later
        year_filter = int(self.year)
        dashboard_stats = self.env['vit.budget_rkap'].get_statistics_sarlog(year_filter)
        grouped_data = {}
        for item in dashboard_stats.get('master_list', []):
            mb = self.env['vit.master_budget'].browse(item.get('id'))
            budgets = self.env['vit.budget_rkap'].search([
                ('master_budget_id', '=', mb.id),
                ('budget_date', '>=', f'{year_filter}-01-01'),
                ('budget_date', '<=', f'{year_filter}-12-31')
            ], limit=1)
            tipe = budgets.tipe_kegiatan if budgets and budgets.tipe_kegiatan else 'Lainnya'
            if tipe not in grouped_data:
                grouped_data[tipe] = []
            row_name = budgets.name if budgets and budgets.name else item.get('name')
            rkap_val = item.get('pagu_izin_prinsip', 0) if item.get('realisasi', 0) > 0 else 0
            # placeholders for the other two columns
            grouped_data[tipe].append({
                'name': row_name,
                'rkap': rkap_val,
                'sisa_pembayaran': 0,
                'izin_terbit': 0,
            })

        # sort categories same order
        sorted_categories = sorted(grouped_data.keys(), key=lambda x: (0 if x == 'biaya' else 1, x))

        for idx, category in enumerate(sorted_categories):
            items = grouped_data[category]
            category_letter = chr(65 + idx)
            display_name = tipe_display_map.get(category, category.upper())

            # category header
            right_fill = category_fill if category == 'biaya' else subtotal_fill
            ws[f'A{row}'].value = category_letter
            ws[f'A{row}'].font = category_font
            ws[f'A{row}'].fill = right_fill
            ws[f'A{row}'].alignment = center_align
            ws[f'A{row}'].border = thin_border

            ws[f'B{row}'].value = display_name
            ws[f'B{row}'].font = category_font
            ws[f'B{row}'].fill = right_fill
            ws[f'B{row}'].alignment = left_align
            ws[f'B{row}'].border = thin_border
            ws.row_dimensions[row].height = 18
            for col in ['C','D','E','F']:
                ws[f'{col}{row}'].fill = right_fill
                ws[f'{col}{row}'].border = thin_border
                if col in ['C','D','E','F']:
                    ws[f'{col}{row}'].number_format = '#,##0'
            row += 1

            # items
            cat_sum_rkap = cat_sum_sisa = cat_sum_izin = 0
            for item in items:
                rkap_val = item.get('rkap', 0)
                sisa_val = item.get('sisa_pembayaran', 0)
                izin_val = item.get('izin_terbit', 0)
                saldo_val = rkap_val - (sisa_val + izin_val)

                ws[f'A{row}'].value = no
                ws[f'B{row}'].value = item['name']
                ws[f'C{row}'].value = rkap_val
                ws[f'D{row}'].value = sisa_val
                ws[f'E{row}'].value = izin_val
                ws[f'F{row}'].value = saldo_val

                all_cols = ['A','B','C','D','E','F']
                for col in all_cols:
                    cell = ws[f'{col}{row}']
                    cell.border = thin_border
                    if col == 'A':
                        cell.alignment = center_align
                    elif col == 'B':
                        cell.alignment = left_align
                    else:
                        cell.alignment = right_align
                        cell.number_format = '#,##0'
                    if right_fill:
                        cell.fill = right_fill
                cat_sum_rkap += rkap_val
                cat_sum_sisa += sisa_val
                cat_sum_izin += izin_val
                row += 1
                no += 1

            # subtotal row
            ws[f'B{row}'].value = f'JUMLAH {display_name}'
            ws[f'B{row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'B{row}'].fill = subtotal_fill
            ws[f'C{row}'].value = cat_sum_rkap
            ws[f'D{row}'].value = cat_sum_sisa
            ws[f'E{row}'].value = cat_sum_izin
            ws[f'F{row}'].value = cat_sum_rkap - (cat_sum_sisa + cat_sum_izin)
            for col in ['A','B','C','D','E','F']:
                ws[f'{col}{row}'].border = thin_border
                if col in ['A','B']:
                    ws[f'{col}{row}'].fill = right_fill
                else:
                    ws[f'{col}{row}'].fill = subtotal_fill
                    ws[f'{col}{row}'].number_format = '#,##0'
            ws.row_dimensions[row].height = 16
            subtotal_rows[category] = row
            row += 1

        # grand total row (sum subtotals)
        ws[f'B{row}'].value = 'TOTAL'
        ws[f'B{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'B{row}'].fill = total_fill
        biaya_row = subtotal_rows.get('biaya')
        investasi_row = subtotal_rows.get('investasi')
        cols = ['C','D','E','F']
        if biaya_row and investasi_row:
            for col in cols:
                ws[f'{col}{row}'].value = f"={col}{biaya_row}+{col}{investasi_row}"
                ws[f'{col}{row}'].number_format = '#,##0'
        elif biaya_row:
            for col in cols:
                ws[f'{col}{row}'].value = f"={col}{biaya_row}"
                ws[f'{col}{row}'].number_format = '#,##0'
        elif investasi_row:
            for col in cols:
                ws[f'{col}{row}'].value = f"={col}{investasi_row}"
                ws[f'{col}{row}'].number_format = '#,##0'
        for col in ['A','B','C','D','E','F']:
            ws[f'{col}{row}'].border = thin_border
            ws[f'{col}{row}'].fill = total_fill
        ws['B'+str(row)].alignment = left_align
        ws.row_dimensions[row].height = 18

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = output.read()
        output.close()

        # Encode and save as attachment
        data_b64 = base64.b64encode(file_content).decode('utf-8')
        attachment = self.env['ir.attachment'].create({
            'name': f'rekap_ketersediaan_sarlog_{self.year}.xlsx',
            'type': 'binary',
            'datas': data_b64,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': 'wizard.rekap.ketersediaan.sarlog',
            'res_id': self.id,
        })

        # Return download URL
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % (attachment.id,),
            'target': 'new',
        }
