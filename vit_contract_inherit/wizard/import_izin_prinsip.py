#!/usr/bin/python
#-*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError
import io
import base64

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    openpyxl = None


class WizardImportIzinPrinsip(models.TransientModel):
    _name = "wizard.import.izin.prinsip"
    _description = "Wizard Import Izin Prinsip dari Excel"

    file_data = fields.Binary(string="File Excel", required=True)
    file_name = fields.Char(string="File Name")

    @api.model
    def create(self, vals):
        if not openpyxl:
            raise UserError(_("Library openpyxl tidak terinstall. Mohon install library openpyxl."))
        return super(WizardImportIzinPrinsip, self).create(vals)

    def action_import_excel(self):
        """
        Import data Izin Prinsip dari file Excel
        Group rows by Name+Budget, create 1 Izin Prinsip per group with multiple job detail items
        """
        if not self.file_data:
            raise UserError(_("Silakan pilih file Excel terlebih dahulu."))

        try:
            # Decode file data
            file_data = base64.b64decode(self.file_data)
            
            # Load workbook
            workbook = load_workbook(io.BytesIO(file_data))
            sheet = workbook.active
            
            # Get header and normalize them to canonical names
            headers = []
            for cell in sheet[1]:
                val = cell.value
                if val is None:
                    headers.append('')
                    continue
                h = str(val).strip()
                h = ' '.join(h.split())
                headers.append(h)

            # Default mapping by column index (A=0, B=1, ...)
            default_index_map = {
                0: 'Budget',
                1: 'Kanwil',
                2: 'Nomor Tracker',
                3: 'Name',
                4: 'Job Izin Prinsip',
            }

            # Create canonical header mapping per column index so we can handle
            # header variations (e.g. extra spaces, different capitalization,
            # or missing headers).
            canonical_map = {}
            for idx, h in enumerate(headers):
                key = (h or '').lower()
                if key in ('kanwil', 'kanwil '):
                    canonical = 'Kanwil'
                elif key in ('budget', 'budget '):
                    canonical = 'Budget'
                elif key.replace(' ', '') in ('nomortracker', 'nomor_tracker', 'nomor tracker'):
                    canonical = 'Nomor Tracker'
                elif key in ('name', 'nama'):
                    canonical = 'Name'
                elif key.replace(' ', '') in ('jobizinprinsip', 'job_izin_prinsip', 'uraianpekerjaan', 'uraian pekerjaan', 'uraian'):
                    canonical = 'Job Izin Prinsip'
                else:
                    canonical = h if h else default_index_map.get(idx, f'col_{idx}')
                canonical_map[idx] = canonical
            
            # Process rows with fill-down
            imported_count = 0
            created_ids = []
            errors = []
            prev_values = [None] * len(headers)
            
            # Group rows by (Name, Budget)
            grouped_rows = {}
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # convert row to list and ensure length matches headers
                    row_list = list(row) if row is not None else []
                    if len(row_list) < len(headers):
                        row_list += [None] * (len(headers) - len(row_list))

                    # Skip completely empty rows
                    if not any([cell for cell in row_list if cell not in (None, '')]):
                        continue

                    # Fill-down: if a cell is empty, use previous non-empty value for that column
                    filled_row = []
                    for col_idx in range(len(headers)):
                        raw_val = row_list[col_idx]
                        use_val = raw_val
                        if isinstance(raw_val, str):
                            if raw_val.strip() == '':
                                use_val = prev_values[col_idx]
                        else:
                            if raw_val is None:
                                use_val = prev_values[col_idx]

                        # update prev_values if we have a new non-empty value
                        if use_val not in (None, ''):
                            prev_values[col_idx] = use_val

                        filled_row.append(use_val)

                    # Create dict from filled row data using canonical headers
                    # (so ' Kanwil', 'kanwil', empty header at column B, etc. are
                    # all mapped to 'Kanwil'). This ensures grouping by Kanwil
                    # works reliably.
                    row_data = {}
                    for col_idx in range(len(headers)):
                        header = canonical_map.get(col_idx, default_index_map.get(col_idx, f'col_{col_idx}'))
                        row_data[header] = filled_row[col_idx]
                    
                    # Validate required fields
                    if not row_data.get('Name'):
                        errors.append(f"Baris {row_idx}: Nama tidak boleh kosong")
                        continue
                    
                    # Group key: (Name, Budget, Kanwil)
                    # Include Kanwil so rows from different regions with the same
                    # Name/Budget aren't merged into a single Izin Prinsip.
                    name_key = (row_data.get('Name') or '')
                    budget_key = (row_data.get('Budget') or '')
                    kanwil_key = (row_data.get('Kanwil') or row_data.get('kanwil') or '')
                    # normalize to stripped strings for consistent grouping
                    try:
                        name_key = str(name_key).strip()
                    except Exception:
                        pass
                    try:
                        budget_key = str(budget_key).strip()
                    except Exception:
                        pass
                    try:
                        kanwil_key = str(kanwil_key).strip()
                    except Exception:
                        pass

                    group_key = (name_key, budget_key, kanwil_key)
                    if group_key not in grouped_rows:
                        grouped_rows[group_key] = []
                    grouped_rows[group_key].append((row_idx, row_data))
                    
                except Exception as e:
                    errors.append(f"Baris {row_idx}: {str(e)}")
            
            # Process each group: 1 Izin Prinsip + multiple Job items
            # grouped_rows keys are (name, budget, kanwil)
            for (name, budget_name, kanwil_name), rows_in_group in grouped_rows.items():
                try:
                    # Get budget (try exact then case-insensitive partial)
                    budget_name = (budget_name or '').strip()
                    budget_id = None
                    budget_rec = None
                    if budget_name:
                        budget_rec = self.env['vit.budget_rkap'].search([('name', '=', budget_name)], limit=1)
                        if not budget_rec:
                            budget_rec = self.env['vit.budget_rkap'].search([('name', 'ilike', budget_name)], limit=1)
                        if not budget_rec:
                            errors.append(f"Budget '{budget_name}' tidak ditemukan (untuk Izin Prinsip: {name})")
                            continue
                        budget_id = budget_rec.id
                    
                    # Get kanwil: prefer kanwil_name from group key, but fallback to
                    # value found in the first row if needed
                    first_row_data = rows_in_group[0][1]
                    kanwil_name = (kanwil_name or first_row_data.get('Kanwil') or first_row_data.get('kanwil') or '').strip()
                    kanwil_id = None
                    kanwil_rec = None
                    if kanwil_name:
                        kanwil_rec = self.env['vit.kanwil'].search([('name', '=', kanwil_name)], limit=1)
                        if not kanwil_rec:
                            kanwil_rec = self.env['vit.kanwil'].search([('name', 'ilike', kanwil_name)], limit=1)
                        if not kanwil_rec:
                            errors.append(f"Kanwil '{kanwil_name}' tidak ditemukan (untuk Izin Prinsip: {name})")
                            continue
                        kanwil_id = kanwil_rec.id
                    
                    # Get stage from first row
                    stage_name = first_row_data.get('Stage', 'Draft')
                    stage = self.env['vit.state_izin_prinsip'].search([('name', '=', stage_name)], limit=1)
                    if not stage:
                        stage = self.env['vit.state_izin_prinsip'].search([], order='sequence asc', limit=1)
                    
                    # Compute expected total_pagu from the Excel rows (sum of contract columns)
                    def _to_number(val):
                        try:
                            if val is None:
                                return 0.0
                            if isinstance(val, (int, float)):
                                return float(val)
                            s = str(val).strip()
                            if s == '':
                                return 0.0
                            # remove common thousand separators
                            s = s.replace(',', '')
                            return float(s)
                        except Exception:
                            return 0.0

                    expected_total_pagu = 0.0
                    for _ri, _rdata in rows_in_group:
                        for col in ('Fisik', 'Perencana', 'Pengawas', 'Pengelola', 'Total', 'Total Pagu'):
                            expected_total_pagu += _to_number(_rdata.get(col))

                    # Create Izin Prinsip (single record per group)
                    izin_prinsip_vals = {
                        'name': name,
                        'nomor_tracker': first_row_data.get('Nomor Tracker') or first_row_data.get('NomorTracker') or first_row_data.get('Nomor_Tracker'),
                        'budget_id': budget_id,
                        'kanwil_id': kanwil_id,
                        'stage_id': stage.id if stage else None,
                        'master_budget_id': budget_rec.master_budget_id.id if budget_rec and budget_rec.master_budget_id else False,
                        'created_by': self.env.user.id,
                        'active': True,
                    }
                    izin_prinsip = self.env['vit.izin_prinsip'].sudo().create(izin_prinsip_vals)
                    created_ids.append(izin_prinsip.id)
                    
                    # Get default kanca for jobs
                    kanca_name = (first_row_data.get('Kanca') or first_row_data.get('kanca') or '').strip()
                    kanca_id = None
                    if kanca_name:
                        kanca_rec = self.env['vit.kanca'].search([('name', '=', kanca_name)], limit=1)
                        if not kanca_rec:
                            kanca_rec = self.env['vit.kanca'].search([('name', 'ilike', kanca_name)], limit=1)
                        if kanca_rec:
                            kanca_id = kanca_rec.id

                    if not kanca_id:
                        if self.env.user.multi_kanca:
                            kanca_id = self.env.user.multi_kanca[0].id
                    
                    # Create Job items for each row in the group
                    for row_idx, row_data in rows_in_group:
                        try:
                            # Get job description from Excel row
                            job_name = None
                            possible_job_keys = [
                                'Job Izin Prinsip', 'Job_Izin_Prinsp', 'Uraian Pekerjaan', 'Uraian_Pekerjaan',
                                'Uraian', 'Uraian_Pekerjaan?', 'Uraian Pekerjaan?'
                            ]
                            for k in possible_job_keys:
                                v = row_data.get(k)
                                if v not in (None, ''):
                                    job_name = str(v).strip()
                                    break

                            if not job_name:
                                job_name = f"Detail {row_idx}"

                            # Extract Kanca name from job_name if it contains "Kanca ..." pattern
                            job_kanca_id = kanca_id  # default to group's kanca
                            job_name_lower = job_name.lower()
                            
                            # Try to find "Kanca [name]" pattern in job_name
                            if 'kanca' in job_name_lower:
                                # Extract word after "Kanca"
                                parts = job_name.split(',')
                                for part in parts:
                                    part_stripped = part.strip()
                                    if part_stripped.lower().startswith('kanca'):
                                        # Remove "Kanca " prefix and get the name
                                        kanca_text = part_stripped[5:].strip()
                                        if kanca_text:
                                            # Search for this kanca
                                            kanca_search = self.env['vit.kanca'].search([('name', 'ilike', kanca_text)], limit=1)
                                            if kanca_search:
                                                job_kanca_id = kanca_search.id
                                            break
                            
                            # If kanwil is mentioned but kanca is not found (job_kanca_id is still the default kanca_id and kanca_id is empty),
                            # create a virtual kanca with name "Kanwil [kanwil_name]"
                            if not job_kanca_id and kanwil_rec:
                                # kanca_id is empty (not found in group), but kanwil_rec exists
                                # Create virtual kanca name: "Kanwil [Kanwil Name]"
                                virtual_kanca_name = f"Kanwil {kanwil_rec.name}"
                                
                                # Try to find existing kanca with this name
                                virtual_kanca = self.env['vit.kanca'].search([('name', '=', virtual_kanca_name)], limit=1)
                                if not virtual_kanca:
                                    # Create new kanca if it doesn't exist
                                    virtual_kanca = self.env['vit.kanca'].sudo().create({
                                        'name': virtual_kanca_name,
                                        'parent_kanwil_id': kanwil_id,
                                    })
                                job_kanca_id = virtual_kanca.id

                            job_vals = {
                                'name': job_name,
                                'izin_prinsip_id': izin_prinsip.id,
                            }
                            if job_kanca_id:
                                job_vals['kanca_id'] = job_kanca_id

                            job_izin_prinsip = self.env['vit.job_izin_prinsip'].sudo().create(job_vals)
                            
                            # Create Izin Prinsip Line items for each contract type with values
                            # Map header names to contract type names
                            contract_type_mapping = {
                                'Fisik': 'Fisik',
                                'Perencana': 'Perencana',
                                'Pengawas': 'Pengawas',
                                'Pengelola': 'Pengelola',
                            }
                            
                            for excel_header, contract_type_name in contract_type_mapping.items():
                                value = row_data.get(excel_header)
                                
                                # Only create line if value exists and is not 0
                                if value and (isinstance(value, (int, float)) and value != 0):
                                    try:
                                        # Find jenis_kontrak by name
                                        jenis_kontrak = self.env['vit.jenis_kontrak'].search(
                                            [('name', '=', contract_type_name)], limit=1
                                        )
                                        
                                        if not jenis_kontrak:
                                            # Try case-insensitive search
                                            jenis_kontrak = self.env['vit.jenis_kontrak'].search(
                                                [('name', 'ilike', contract_type_name)], limit=1
                                            )
                                        
                                        if jenis_kontrak:
                                            line_vals = {
                                                'name': contract_type_name,
                                                'pagu': value,
                                                'jenis_kontrak_id': jenis_kontrak.id,
                                                'job_izin_prinsip_id': job_izin_prinsip.id,
                                            }
                                            if job_kanca_id:
                                                line_vals['kanca_id'] = job_kanca_id
                                            if kanwil_id:
                                                line_vals['kanwil_id'] = kanwil_id
                                            
                                            self.env['vit.izin_prinsip_line'].sudo().create(line_vals)
                                    except Exception as e:
                                        errors.append(f"Baris {row_idx} (Izin Prinsip Line {contract_type_name}): {str(e)}")
                        except Exception as e:
                            errors.append(f"Baris {row_idx} (Job detail): {str(e)}")
                    
                    imported_count += 1
                    
                except Exception as e:
                    # Add more context for budget overrun errors
                    try:
                        budget_rem = budget_rec.remaining if budget_rec else 'N/A'
                    except Exception:
                        budget_rem = 'N/A'
                    errors.append(
                        f"Izin Prinsip '{name}' (Budget: {budget_name}, Kanwil: {kanwil_name}): {str(e)} -- Expected Total Pagu: {expected_total_pagu:.2f}, Budget Remaining: {budget_rem}"
                    )
            
            # If we created records, open a list view showing them so user can
            # immediately see imported items even if global record rules hide others
            if imported_count > 0:
                action = {
                    'type': 'ir.actions.act_window',
                    'name': _('Imported Izin Prinsip'),
                    'res_model': 'vit.izin_prinsip',
                    'view_mode': 'tree,form',
                    'domain': [('id', 'in', created_ids)],
                    'context': dict(self.env.context, active_test=False),
                }
                return action

            # otherwise show errors
            message = f"Berhasil import {imported_count} Izin Prinsip"
            if errors:
                message += f"\n\nError yang ditemukan:\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    message += f"\n... dan {len(errors) - 10} error lainnya"

            raise UserError(_(message))
            
        except UserError:
            raise
        except Exception as e:
            raise UserError(_("Error saat membaca file Excel: %s") % str(e))
